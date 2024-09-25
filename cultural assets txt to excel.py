"""
20240925 愛連網作業：文化資產圖資txt轉到excel
逐行讀取txt檔內資料再寫入excel
"""

from openpyxl import load_workbook

# 設定excel匯出檔案及名稱
fileRoute = 'Cultural_to_excel_Test.xlsx'
sheetName = 'sheet1'
head = 1  # 有一列表頭

# excel欄位設定
workFile = load_workbook(filename=fileRoute, read_only=False)
workSheet = workFile[sheetName]
col = list(workSheet.columns)
row = head + 1

f = open('cultural.txt', 'r', encoding="utf-8")  # 讀取的txt檔
while True:  # txt檔逐行讀取
    line = f.readline()
    if line:
        if 'caseId' in line:
            # print(line[15:])
            workSheet.cell(row, 1).value = line[15:]
        if 'caseName' in line:
            # print(line[17:])
            workSheet.cell(row, 2).value = line[17:]
        if 'assetsClassifyName' in line:
            # print(line[27:])
            workSheet.cell(row, 3).value = line[27:]
        if 'assetsTypeName' in line:
            # print(line[23:])
            workSheet.cell(row, 4).value = line[23:]
        if 'representImage' in line:
            # print(line[23:])
            workSheet.cell(row, 5).value = line[23:]
        if 'cityName' in line:
            # print(line[17:])
            workSheet.cell(row, 6).value = line[17:]
        if 'lat_boch' in line:
            # print(line[17:])
            workSheet.cell(row, 7).value = line[17:]
        if 'lon_boch' in line:
            # print(line[17:])
            workSheet.cell(row, 8).value = line[17:]
        if 'latitude' in line:
            # print(line[17:])
            workSheet.cell(row, 9).value = line[17:]
        if 'longitude' in line:
            # print(line[18:])
            workSheet.cell(row, 10).value = line[18:]
        if 'Year' in line:
            # print(line[12:])
            workSheet.cell(row, 11).value = line[12:]
            row += 1  # 換行
    else:
        break  # 除上述if條件設定外跳出迴圈
workFile.save(fileRoute)  # excel存檔
f.close()  # 關閉txt檔
print("Finish!")
